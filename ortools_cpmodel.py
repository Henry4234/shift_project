from ortools.sat.python import cp_model
import pandas as pd
import openpyxl
from openpyxl import workbook
from openpyxl.utils import get_column_letter,column_index_from_string



def schedule_shift_planning(num_days=31, num_employees=8):
    # 創建模型
    model = cp_model.CpModel()

    # 參數
    shifts = [1, 2, 3]  # 白班、小夜班、大夜班
    max_work_days_per_two_weeks = 12  # 兩周最多工作天數
    min_rest_days_per_two_weeks = 2   # 兩周最少休息天數

    # 定義需求人力 (使用用戶提供的需求列表)
    need = [
        [1, 1, 3], [1, 2, 2], [1, 3, 1], [2, 1, 3], [2, 2, 2], [2, 3, 1],
        [3, 1, 3], [3, 2, 2], [3, 3, 1], [4, 1, 2], [4, 2, 1], [4, 3, 1],
        [5, 1, 1], [5, 2, 1], [5, 3, 1], [6, 1, 3], [6, 2, 2], [6, 3, 1],
        [7, 1, 3], [7, 2, 2], [7, 3, 1], [8, 1, 3], [8, 2, 2], [8, 3, 1],
        [9, 1, 3], [9, 2, 2], [9, 3, 1], [10, 1, 3], [10, 2, 2], [10, 3, 1],
        [11, 1, 2], [11, 2, 1], [11, 3, 1], [12, 1, 1], [12, 2, 1], [12, 3, 1],
        [13, 1, 3], [13, 2, 2], [13, 3, 1], [14, 1, 3], [14, 2, 2], [14, 3, 1],
        [15, 1, 3], [15, 2, 2], [15, 3, 1], [16, 1, 3], [16, 2, 2], [16, 3, 1],
        [17, 1, 3], [17, 2, 2], [17, 3, 1], [18, 1, 2], [18, 2, 1], [18, 3, 1],
        [19, 1, 1], [19, 2, 1], [19, 3, 1], [20, 1, 3], [20, 2, 2], [20, 3, 1],
        [21, 1, 3], [21, 2, 2], [21, 3, 1], [22, 1, 3], [22, 2, 2], [22, 3, 1],
        [23, 1, 3], [23, 2, 2], [23, 3, 1], [24, 1, 3], [24, 2, 2], [24, 3, 1],
        [25, 1, 2], [25, 2, 1], [25, 3, 1], [26, 1, 1], [26, 2, 1], [26, 3, 1],
        [27, 1, 3], [27, 2, 2], [27, 3, 1], [28, 1, 3], [28, 2, 2], [28, 3, 1],
        [29, 1, 3], [29, 2, 2], [29, 3, 1], [30, 1, 3], [30, 2, 2], [30, 3, 1],
        [31, 1, 3], [31, 2, 2], [31, 3, 1]
    ]

    # 決策變數
    x = {}
    for d in range(num_days):
        for s in shifts:
            for e in range(num_employees):
                x[d, s, e] = model.NewBoolVar(f'x[{d},{s},{e}]')  # 是否在 d 天 s 班 e 員工排班

    # 目標式: 最小化員工工作天數差異
    total_work = [model.NewIntVar(0, num_days, f'total_work[{e}]') for e in range(num_employees)]
    for e in range(num_employees):
        model.Add(total_work[e] == sum(x[d, s, e] for d in range(num_days) for s in shifts))
    max_work = model.NewIntVar(0, num_days, 'max_work')
    min_work = model.NewIntVar(0, num_days, 'min_work')
    model.AddMaxEquality(max_work, total_work)
    model.AddMinEquality(min_work, total_work)
    model.Minimize(max_work - min_work)

    # 限制式

    # 每天每班決定需求人數 (使用 need 數據)
    for d in range(num_days):
        for s in shifts:
            required_staff = next((n[2] for n in need if n[0] == d + 1 and n[1] == s), 0)
            model.Add(sum(x[d, s, e] for e in range(num_employees)) == required_staff)

    # 每位員工每天最多工作一個班次
    for d in range(num_days):
        for e in range(num_employees):
            model.Add(sum(x[d, s, e] for s in shifts) <= 1)

    # 禁止大夜班後接白班 & 小夜班
    for d in range(num_days - 1):  # 從第1天到倒數第2天
        for e in range(num_employees):
            model.Add(x[d, 3, e] + x[d + 1, 1, e] <= 1)
            model.Add(x[d, 3, e] + x[d + 1, 2, e] <= 1)

    # 禁止小夜班後接白班
    for d in range(num_days - 1):  # 從第1天到倒數第2天
        for e in range(num_employees):
            model.Add(x[d, 2, e] + x[d + 1, 1, e] <= 1)

    # 每兩周（14天）內至少休息2天
    for e in range(num_employees):
        for d in range(num_days - 13):  # 滑動窗口範圍
            model.Add(sum(x[d + i, s, e] for i in range(14) for s in shifts) <= max_work_days_per_two_weeks)

    # 解決模型
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    # 結果輸出
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        # print('解決方案:')
        schedule_data = []  # 用於存儲 DataFrame 資料
        for d in range(num_days):
            # print(f'日期 {d + 1}:')
            for s in shifts:
                assigned = []
                for e in range(num_employees):
                    if solver.Value(x[d, s, e]) == 1:
                        assigned.append(e)
                        schedule_data.append({"日期": d + 1, "班次": s, "員工": e})
                # print(f'  班次 {s}: {assigned}')
        # print('每位員工的工作天數:')
        # for e in range(num_employees):
        #     print(f'  員工 {e}: {solver.Value(total_work[e])} 天')

        # 將排班結果匯出為 DataFrame
        global schedule_df
        schedule_df = pd.DataFrame(schedule_data,columns=["日期", "班次", "員工"])
        schedule_df.to_excel("schedule_output.xlsx", index=False)
        print(schedule_df)
    else:
        print('沒有可行解決方案')

# 執行模型
schedule_shift_planning(num_days=31, num_employees=8)


##-------------------------------------------openpyxl------------------------------------------##
# 讀取文件
file_name = "./schedule_output.xlsx"

WB = openpyxl.open(file_name)
#建立新的工作表"BI"
WB.create_sheet("BI")
#插入第一列"日期"
days = ["日期"] + [d for d in range(1,32)]
ws = WB["BI"]
ws.append(days)
#插入第一欄"員工"
for i in range(0,8):
    ws.append([i])
# 將 `schedule_df` 轉為字典，便於快速查詢

# 定義班別與代碼的映射
shift_code_mapping = {
    "白班8-16點": "A",
    "小夜班16-24點": "B",
    "大夜班0-8點": "C",
    1:"A",
    2: "B",
    3: "C",
}

schedule_df["班次"] = schedule_df["班次"].replace(shift_code_mapping)
# 字典結構：{(日期, 員工): 班次}
schedule_dict = {
    (row["日期"], row["員工"]): row["班次"]
    for _, row in schedule_df.iterrows()
}
# print(schedule_dict)

# 填入班次資料
employees = [i for i in range(8)]
for row_index, emp in enumerate(employees, start=2):  # 員工從第二行開始
    for col_index, day in enumerate(days[1:], start=2):  # 日期從第二列開始 (跳過 "日期")
        # 查詢當天是否有該員工的班次
        if (day, emp) in schedule_dict:
            ws.cell(row=row_index, column=col_index).value = schedule_dict[(day, emp)]

ws.append([])
count_lst = ["C_A","C_B","C_C"]
for row in range(11,14):
    ws["A" + str(row)] = count_lst[row-11]
for col in range(2,33):
    char = get_column_letter(col)
    for row in range(11,14):
        if row == 11:
            ws[char + str(row)] = '=COUNTIF(%s2:%s9,"A")'%(char,char)  #=COUNTIF(B2:B9,"A")
        elif row == 12:
            ws[char + str(row)] = '=COUNTIF(%s2:%s9,"B")'%(char,char)  #=COUNTIF(B2:B9,"B")
        else:
            ws[char + str(row)] = '=COUNTIF(%s2:%s9,"C")'%(char,char)  #=COUNTIF(B2:B9,"C")
WB.save(file_name)
