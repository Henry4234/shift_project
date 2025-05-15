# 使用 MindOpt 建立的排班模型 (Python 實現)
import pandas as pd
from ortools.linear_solver import pywraplp
import openpyxl
from openpyxl import workbook
from openpyxl.utils import get_column_letter,column_index_from_string
# 載入輸入數據
# schedule_df = pd.read_csv("班次.csv")
# num_demand_df = pd.read_csv("需求人数.csv")

# num_demand_long_df = pd.read_csv("需求人数-长列表.csv")
data = [[1,	"白班8-16點"	,3],
[1,	"小夜班16-24點"	,2],
[1,	"大夜班0-8點"	,1],
[2,	"白班8-16點"	,3],
[2,	"小夜班16-24點"	,2],
[2,	"大夜班0-8點"	,1],
[3,	"白班8-16點"	,3],
[3,	"小夜班16-24點"	,2],
[3,	"大夜班0-8點"	,1],
[4,	"白班8-16點"	,2],
[4,	"小夜班16-24點"	,1],
[4,	"大夜班0-8點"	,1],
[5,	"白班8-16點"	,1],
[5,	"小夜班16-24點"	,1],
[5,	"大夜班0-8點"	,1],
[6,	"白班8-16點"	,3],
[6,	"小夜班16-24點"	,2],
[6,	"大夜班0-8點"	,1],
[7,	"白班8-16點"	,3],
[7,	"小夜班16-24點"	,2],
[7,	"大夜班0-8點"	,1],
[8,	"白班8-16點"	,3],
[8,	"小夜班16-24點"	,2],
[8,	"大夜班0-8點"	,1],
[9,	"白班8-16點"	,3],
[9,	"小夜班16-24點"	,2],
[9,	"大夜班0-8點"	,1],
[10,	"白班8-16點"	,3],
[10,	"小夜班16-24點"	,2],
[10,	"大夜班0-8點"	,1],
[11,	"白班8-16點"	,2],
[11,	"小夜班16-24點"	,1],
[11,	"大夜班0-8點"	,1],
[12,	"白班8-16點"	,1],
[12,	"小夜班16-24點"	,1],
[12,	"大夜班0-8點"	,1],
[13,	"白班8-16點"	,3],
[13,	"小夜班16-24點"	,2],
[13,	"大夜班0-8點"	,1],
[14,	"白班8-16點"	,3],
[14,	"小夜班16-24點"	,2],
[14,	"大夜班0-8點"	,1],
[15,	"白班8-16點"	,3],
[15,	"小夜班16-24點"	,2],
[15,	"大夜班0-8點"	,1],
[16,	"白班8-16點"	,3],
[16,	"小夜班16-24點"	,2],
[16,	"大夜班0-8點"	,1],
[17,	"白班8-16點"	,3],
[17,	"小夜班16-24點"	,2],
[17,	"大夜班0-8點"	,1],
[18,	"白班8-16點"	,2],
[18,	"小夜班16-24點"	,1],
[18,	"大夜班0-8點"	,1],
[19,	"白班8-16點"	,1],
[19,	"小夜班16-24點"	,1],
[19,	"大夜班0-8點"	,1],
[20,	"白班8-16點"	,3],
[20,	"小夜班16-24點"	,2],
[20,	"大夜班0-8點"	,1],
[21,	"白班8-16點"	,3],
[21,	"小夜班16-24點"	,2],
[21,	"大夜班0-8點"	,1],
[22,	"白班8-16點"	,3],
[22,	"小夜班16-24點"	,2],
[22,	"大夜班0-8點"	,1],
[23,	"白班8-16點"	,3],
[23,	"小夜班16-24點"	,2],
[23,	"大夜班0-8點"	,1],
[24,	"白班8-16點"	,3],
[24,	"小夜班16-24點"	,2],
[24,	"大夜班0-8點"	,1],
[25,	"白班8-16點"	,2],
[25,	"小夜班16-24點"	,1],
[25,	"大夜班0-8點"	,1],
[26,	"白班8-16點"	,1],
[26,	"小夜班16-24點"	,1],
[26,	"大夜班0-8點"	,1],
[27,	"白班8-16點"	,3],
[27,	"小夜班16-24點"	,2],
[27,	"大夜班0-8點"	,1],
[28,	"白班8-16點"	,3],
[28,	"小夜班16-24點"	,2],
[28,	"大夜班0-8點"	,1],
[29,	"白班8-16點"	,3],
[29,	"小夜班16-24點"	,2],
[29,	"大夜班0-8點"	,1],
[30,	"白班8-16點"	,3],
[30,	"小夜班16-24點"	,2],
[30,	"大夜班0-8點"	,1],
[31,	"白班8-16點"	,3],
[31,	"小夜班16-24點"	,2],
[31,	"大夜班0-8點"	,1]]



num_demand_long_df = pd.DataFrame(data,columns=["日期", "班次", "人数"])
# print(num_demand_long_df)
# print(num_demand_long_df["日期"])
# 參數設定
# schedules = schedule_df["班次"].tolist()  # 班次列表
schedules = ["白班8-16點","小夜班16-24點","大夜班0-8點"]
# days = num_demand_df["日期"].tolist()  # 日期列表
days = [d for d in range(1,32)]
max_day = max(days)  # 最大日期
# total_slots = num_demand_long_df["人数"].sum()  # 總排班需求數量
total_slots = 6

# 員工數量估算
estimated_employees = int(total_slots / 5) + 1  # 預估需要的員工數量
employees = [f"員工 {i+1}號" for i in range(0,8)]  # 員工列表
matrix= [[13,	4,	3],
[10,	9,	0],
[9,	8,	0],
[11,	6,	3],
[10,	6,	4],
[11,	9,	0],
[3,	0,	17],
[7,	9,	4]] ##員工天數
employees_days = {employee: matrix[i] for i, employee in enumerate(employees)}


# employees = [y for y in range(1,9)]
# 建立求解器
# solver = pywraplp.Solver.CreateSolver("CBC")
solver = pywraplp.Solver.CreateSolver("GLOP")
if not solver:
    raise Exception("找不到求解器！")

# 定義決策變數
x = {}
for day in days:
    for schedule in schedules:
        for emp in employees:
            x[day, schedule, emp] = solver.BoolVar(f"x[{day},{schedule},{emp}]")

# 約束條件
# 1. 滿足每天每個班次的需求人數
for day in days:
    for schedule in schedules:
        solver.Add(
            sum(x[day, schedule, emp] for emp in employees) >= num_demand_long_df[(num_demand_long_df["日期"] == day) & (num_demand_long_df["班次"] == schedule)]["人数"].iloc[0]
        )

# 2. 每位員工每天最多只能排一個班次
for day in days:
    for emp in employees:
        solver.Add(sum(x[day, schedule, emp] for schedule in schedules) <= 1)

# # 3-1. 禁止大夜班後接白班 & 小夜班後接白班
# for emp in employees:
#     for day in range(1, max_day):
#         if day + 1 <= max_day:
#             # slack_var_upcommingday = solver.NumVar(0, 1, f"slack_night_to_morning[{day},{emp}]")
#             # slack_var_upcommingnight = solver.NumVar(0, 1, f"slack_small_night_to_morning[{day},{emp}]")
#             solver.Add(
#                 x[day, "大夜班0-8點", emp] + x[day + 1, "白班8-16點", emp] <= 1 
#             )
#             solver.Add(
#                 x[day, "小夜班16-24點", emp] + x[day + 1, "白班8-16點", emp] <= 1
#             )

# 4. 每位員工不得連續上班超過 7 天
for emp in employees:
    for start_day in range(1, max_day - 7 + 2):  # 設定窗口起點範圍，確保窗口不超出總日期
        # slack_var_7days = solver.NumVar(0, 7, f"slack_continuous[{emp},{start_day}]")
        solver.Add(
            sum(
                sum(x[start_day + offset, schedule, emp] for schedule in schedules)
                for offset in range(7)  # 檢查連續 7 天
            ) <= 7   # 加入 slack 變數允許部分違反
        )

# 6. 員工 7 號希望大夜班連續 6 天後休息 2 天
# target_emp = "員工 7號"
# for start_day in range(1, max_day - 7 + 1):  # 確保窗口不超出總日期
#     # 計算連續 6 天的大夜班數量
#     consecutive_night_shifts = sum(x[start_day + offset, "大夜班0-8點", target_emp] for offset in range(6))
#     slack_var_night_shifts = solver.NumVar(0, 6, f"slack_night_shifts[{start_day}]")
#     solver.Add(consecutive_night_shifts + slack_var_night_shifts  == 6)
#     # 計算接下來 2 天的休息情況
#     consecutive_rest_days = sum(
#         sum(x[start_day + 6 + rest_day, schedule, target_emp] for schedule in schedules)
#         for rest_day in range(2)
#     )
#     # 限制條件：連續 6 天的大夜班且接下來 2 天休息
#     slack_var_night_rest = solver.NumVar(0, 2, f"slack_night_rest[{start_day}]")
#     solver.Add(consecutive_rest_days + slack_var_night_rest == 0 )

# # 5. 滿足每位員工的班次需求總數
# slack_vars = {}
# for emp, (white_shifts, small_night_shifts, big_night_shifts) in employees_days.items():
#     # 白班需求
#     slack_white = solver.NumVar(0, white_shifts, f"slack_white[{emp}]")
#     solver.Add(
#         sum(x[day, "白班8-16點", emp] for day in days) >= white_shifts - slack_white
#     )
#     # 小夜班需求
#     slack_small_night = solver.NumVar(0, small_night_shifts, f"slack_small_night[{emp}]")
#     solver.Add(
#         sum(x[day, "小夜班16-24點", emp] for day in days) >= small_night_shifts - slack_small_night
#     )
#     # 大夜班需求
#     slack_big_night = solver.NumVar(0, big_night_shifts, f"slack_big_night[{emp}]")
#     solver.Add(
#         sum(x[day, "大夜班0-8點", emp] for day in days) >= big_night_shifts - slack_big_night
#     )
#     slack_vars[emp] = [slack_white, slack_small_night, slack_big_night]




# 目標函數：最小化分配的班次總數
objective = solver.Objective()
# 最小化班次
for day, schedule, emp in x:
    objective.SetCoefficient(x[day, schedule, emp], 1)
# for emp in employees:
#     for day in range(1, max_day):
#         if day + 1 <= max_day:
#             objective.SetCoefficient(slack_var_upcommingday, 30)  # 設定權重
#             objective.SetCoefficient(slack_var_upcommingnight, 30)  # 設定權重


# 將 slack 加入到目標函數中（權重介於高於條件 5，低於條件 1）
# for start_day in range(1, max_day - 7 + 1):  
#     objective.SetCoefficient(slack_var_night_rest, 0)
#     objective.SetCoefficient(slack_var_night_shifts, 0)

# for emp in employees:
#     slack_white, slack_small_night, slack_big_night = slack_vars[emp]
#     objective.SetCoefficient(slack_white, 20)  
#     objective.SetCoefficient(slack_small_night, 20)
#     objective.SetCoefficient(slack_big_night, 20)

objective.SetMinimization()

# 求解
status = solver.Solve()

if status == pywraplp.Solver.OPTIMAL:
    print("找到最佳解！")
    schedule_data = []
    for day in days:
        for schedule in schedules:
            for emp in employees:
                if x[day, schedule, emp].solution_value() > 0.5:
                    schedule_data.append([day, schedule, emp])
    schedule_df = pd.DataFrame(schedule_data, columns=["日期", "班次", "員工"])
    schedule_df.to_excel("schedule_output.xlsx", index=False)
    print("排班結果已保存至 schedule_output.xlsx")
else:
    print("找不到可行解！")


##-------------------------------------------openpyxl------------------------------------------##
# 讀取文件
file_name = "./schedule_output.xlsx"

WB = openpyxl.open(file_name)
#建立新的工作表"BI"
WB.create_sheet("BI")
#插入第一列"日期"
days.insert(0,"日期")
ws = WB["BI"]
ws.append(days)
#插入第一欄"員工"
for i in range(0,len(employees)):
    ws.append([employees[i]])
# 將 `schedule_df` 轉為字典，便於快速查詢

# 定義班別與代碼的映射
shift_code_mapping = {
    "白班8-16點": "A",
    "小夜班16-24點": "B",
    "大夜班0-8點": "C"
}
schedule_df["班次"] = schedule_df["班次"].replace(shift_code_mapping)
# 字典結構：{(日期, 員工): 班次}
schedule_dict = {
    (row["日期"], row["員工"]): row["班次"]
    for _, row in schedule_df.iterrows()
}
# print(schedule_dict)

# 填入班次資料
for row_index, emp in enumerate(employees, start=2):  # 員工從第二行開始
    for col_index, day in enumerate(days[1:], start=2):  # 日期從第二列開始 (跳過 "日期")
        # 查詢當天是否有該員工的班次
        if (day, emp) in schedule_dict:
            ws.cell(row=row_index, column=col_index).value = schedule_dict[(day, emp)]

ws.append([])
count_lst = ["C_A","C_B","C_C"]
for row in range(11,14):
    ws["A" + str(row)] = count_lst[row-11]
for col in range(2,33):
    char = get_column_letter(col)
    for row in range(11,14):
        if row == 11:
            ws[char + str(row)] = '=COUNTIF(%s2:%s9,"A")'%(char,char)  #=COUNTIF(B2:B9,"A")
        elif row == 12:
            ws[char + str(row)] = '=COUNTIF(%s2:%s9,"B")'%(char,char)  #=COUNTIF(B2:B9,"B")
        else:
            ws[char + str(row)] = '=COUNTIF(%s2:%s9,"C")'%(char,char)  #=COUNTIF(B2:B9,"C")
WB.save(file_name)
